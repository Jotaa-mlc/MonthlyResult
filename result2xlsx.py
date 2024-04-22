from month_result import MonthResult
from bills_plan import BillsPlan
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side
import settings

thick = Side(style='thick', color='000000')
thin = Side(style='thin', color='000000')

title_font = Font(size = 13, bold = True, color = 'FFFFFF')
title_fill = PatternFill("solid", fgColor = '000000')
title_border = Border(top=thick, bottom=thick, left=thick, right=thick)

account_font = Font(size=12, bold=True)
account_border = title_border

sub_account_font = Font(size=11)
sub_account_border = Border(top=thin, bottom=thin, left=thin, right=thin)

number_format = '#,##0.00;[Red]#,##0.00'

def GenerateHeaders(export_file:Worksheet) -> None:
    
    export_file['A1'].value = 'Cód. Conta'
    export_file['B1'].value = 'Descrição'
    
    titles_cells = ['A1', 'B1']
    for cell in titles_cells:
        export_file[cell].font = title_font
        export_file[cell].fill = title_fill
        export_file[cell].border = title_border
    
    model = BillsPlan()
    
    header_row = 2
    for account_index in range(len(model.accounts)):
        export_file["A" + str(header_row)].value = account_index + 1
        export_file["B" + str(header_row)].value = model.accounts[account_index].description
        
        accounts_cells = [col + str(header_row) for col in ['A', 'B']]
        for account_cell in accounts_cells:
            export_file[account_cell].font = account_font
            export_file[account_cell].border = account_border
                    
        header_row += 1
        for sub_account_index in range(len(model.accounts[account_index].sub_accounts)):
            export_file["A" + str(header_row)].value = sub_account_index + 1
            export_file["B" + str(header_row)].value = model.accounts[account_index].sub_accounts[sub_account_index].description
            
            sub_accounts_cells = [col + str(header_row) for col in ['A', 'B']]
            for sub_account_cell in sub_accounts_cells:
                export_file[sub_account_cell].font = sub_account_font
                export_file[sub_account_cell].border = sub_account_border
                
            header_row += 1
    
    export_file['A' + str(header_row)].value = 'Resultado'
    export_file['A' + str(header_row + 1)].value = 'Estoque - Custo'
    
    for row in range(header_row, header_row + 2):
        export_file.merge_cells('A' + str(row) + ':B' + str(row))
        export_file['A' + str(row)].font = title_font
        export_file['A' + str(row)].fill = title_fill
        export_file['A' + str(row)].border = title_border
    
def Export2XLSX(results:list[MonthResult], export_file:Worksheet) -> None:
    result_colum = 3
    for result in results:
        row = 1
        export_file.cell(row, result_colum).value = '{:02d}-{:04d}'.format(result.month, result.year)
        export_file.cell(row, result_colum).font = title_font
        export_file.cell(row, result_colum).fill = title_fill
        export_file.cell(row, result_colum).border = title_border
        
        row += 1
        for account_index in range(len(result.bills_plan.accounts)):
            export_file.cell(row, result_colum).value = result.bills_plan.accounts[account_index].value
            export_file.cell(row, result_colum).number_format = number_format
            export_file.cell(row, result_colum).font = account_font
            export_file.cell(row, result_colum).border = account_border
                        
            row += 1
            for sub_account_index in range(len(result.bills_plan.accounts[account_index].sub_accounts)):
                export_file.cell(row, result_colum).value = result.bills_plan.accounts[account_index].sub_accounts[sub_account_index].value
                export_file.cell(row, result_colum).number_format = number_format
                export_file.cell(row, result_colum).font = sub_account_font
                export_file.cell(row, result_colum).border = sub_account_border
                    
                row += 1
        
        export_file.cell(row, result_colum).value = result.result
        export_file.cell(row+1, result_colum).value = result.stock_cost
        for final in range(row, row+2):
            export_file.cell(final, result_colum).number_format = number_format
            export_file.cell(final, result_colum).font = title_font
            export_file.cell(final, result_colum).fill = title_fill
            export_file.cell(final, result_colum).border = title_border
        
        result_colum += 1
        
        
def AutofitWS(ws:Worksheet) -> None:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width