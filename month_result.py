import settings
import common
import datetime
from bills_plan import BillsPlan
from openpyxl import load_workbook

class MonthResult:
    def __init__(self, _month:str, _year:str) -> None:
        self.month = _month
        self.year = _year
        self.result = 0
        self.stock_cost = 0
        self.bills_plan = BillsPlan()
    
    def CalcResult(self) -> None:
        for account in self.bills_plan.accounts:
            self.result += account.value
    
    def CalcStockCost(self) -> None:
        file_path = common.StockCCVFilePath(self.month, self.year)
        stock_ws = common.LoadSheet(file_path)
        
        if stock_ws == None:
            print(f"ERRO: Não foi possível carregar o arquivo {file_path}")
            return None
        
        for product_row in range(1, stock_ws.max_row):
            amount = stock_ws["C" + str(product_row)].value
            cost = stock_ws["E" + str(product_row)].value
            self.stock_cost += amount * cost
    
    def CalcReceitas(self) -> None:
        document_ws = common.LoadSheet(settings.document_file)

        if document_ws == None:
            print(f"ERRO: Não foi possível carregar o arquivo {settings.document_file}")
            return None
        
        for document_row in range(1, document_ws.max_row):
            is_pedido = document_ws["A" + str(document_row)].value == "Pedido   Saída"
            is_fechado = document_ws["C" + str(document_row)].value == "Fechado"
            is_from_month = common.IsFromMonth(document_ws["I" + str(document_row)].value, self.month, self.year) if is_fechado else False
            
            if is_pedido and is_fechado and is_from_month:
                self.bills_plan.accounts[0].sub_accounts[0].value += document_ws["V" + str(document_row)].value #DINHEIRO
                self.bills_plan.accounts[0].sub_accounts[1].value += document_ws["W" + str(document_row)].value #CHEQUE
                self.bills_plan.accounts[0].sub_accounts[5].value += document_ws["X" + str(document_row)].value #TRANSFERENCIA / PIX
                self.bills_plan.accounts[0].sub_accounts[2].value += document_ws["Y" + str(document_row)].value #CARTAO
                self.bills_plan.accounts[0].sub_accounts[3].value += document_ws["Z" + str(document_row)].value #FATURADO
                self.bills_plan.accounts[0].sub_accounts[3].value += document_ws["AA" + str(document_row)].value #FINANCEIRA
        
        self.bills_plan.accounts[0].CalculateAccount()
    
    def CalcDespesas(self) -> None:
        #CALCULANDO DESPESAS DA TABELA DE CONTAS A PAGAR
        payments_ws = common.LoadSheet(settings.payment_file)
        
        if payments_ws == None:
            print(f"ERRO: Não foi possível carregar o arquivo {settings.payment_file}")
            return None
        
        for payment_row in range(1, payments_ws.max_row):
            is_quitado = payments_ws["H" + str(payment_row)].value == "Quitada"
            is_2plano_contas = payments_ws["M" + str(payment_row)].value == "True"
            is_from_month = common.IsFromMonth(payments_ws["P" + str(payment_row)].value, self.month, self.year) if is_quitado else False
            
            if is_quitado and is_from_month and is_2plano_contas:
                cod_conta = payments_ws["N" + str(payment_row)].value - 1
                cod_subconta = payments_ws["O" + str(payment_row)].value - 1
                
                self.bills_plan.accounts[cod_conta].sub_accounts[cod_subconta].value -= payments_ws["S" + str(payment_row)].value
        
        #CALCULANDO DESPESAS DA TABELA DE LANCAMENTOS LIVRO CAIXA
        cashier_ws = common.LoadSheet(settings.cashier_file)
        
        if cashier_ws == None:
            print(f"ERRO: Não foi possível carregar o arquivo {settings.cashier_file}")
            return None
        
        for payment_row in range(1, cashier_ws.max_row):
            is_2plano_contas = cashier_ws["H" + str(payment_row)].value == "True"
            is_from_month = common.IsFromMonth(cashier_ws["C" + str(payment_row)].value, self.month, self.year) if is_2plano_contas else False
            
            if is_2plano_contas and is_from_month:
                is_saida = cashier_ws["D" + str(payment_row)].value == "Saída"
                cod_conta = cashier_ws["I" + str(payment_row)].value - 1
                cod_subconta = cashier_ws["J" + str(payment_row)].value - 1
                
                payment_value = cashier_ws["G" + str(payment_row)].value
                payment_value = -1 * payment_value if is_saida else payment_value
                                
                self.bills_plan.accounts[cod_conta].sub_accounts[cod_subconta].value -= cashier_ws["G" + str(payment_row)].value
            
        for account in self.bills_plan.accounts:
            account.CalculateAccount()
            
        
    def PrintBillsPlan(self) -> None:
        for account in self.bills_plan.accounts:
            account.PrintAccount()
            
        print(common.FormatPrint('Resultado', self.result, settings.print_size + 4))
        print(common.FormatPrint('Estoque - Custo', self.stock_cost, settings.print_size + 4))
        