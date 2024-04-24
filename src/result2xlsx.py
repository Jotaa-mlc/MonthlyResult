from month_result import MonthResult
from bills_plan import BillsPlan
from settings import export_sheets
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side

export_sheets = export_sheets()

thick = Side(style='thick', color='000000')
thin = Side(style='thin', color='000000')

title_font = Font(size = 13, bold = True, color = 'FFFFFF')
title_fill = PatternFill("solid", fgColor = '000000')
title_border = Border(top=thick, bottom=thick, left=thick, right=thick)

account_font = Font(size=12, bold=True)
account_border = title_border

sub_account_font = Font(size=11)
sub_account_border = Border(top=thin, bottom=thin, left=thin, right=thin)

number_format = '#,##0.00;[Red]#,##0.00'
percentage_format = '0.00%;[Red]0.00%'

def FormatCell(cell:Cell, format_type:int, percentage:bool = False) -> None:
    """Format Cell to correspondent style
    
    Args:
        cell (Cell): Cell to be formated
        format_type (int): 0 - Title; 1 - Account; 2 - Sub Account
        percentage (bool, False): Format as Percentage. Defaults to Number.
    """
    match format_type:
        case 0: #TITLE
            cell.font = title_font
            cell.fill = title_fill
            cell.border = title_border
        case 1: #ACCOUNT
            cell.font = account_font
            cell.border = account_border
        case 2: #SUB_ACCOUNT
            cell.font = sub_account_font
            cell.border = sub_account_border
    
    if type(cell.value) == float:
        cell.number_format = number_format if not percentage else percentage_format

def GenerateHeaders(export_ws:Worksheet, full_headers:bool = True) -> None:
    export_ws['A1'].value = 'Cód. Conta'
    export_ws['B1'].value = 'Descrição'
    model = BillsPlan()
    
    titles_cells = [export_ws['A1'], export_ws['B1']]
    for cell in titles_cells:
        FormatCell(cell, 0)
    
    header_row = 2
    for account in model.accounts:
        account_index = model.accounts.index(account)
        if full_headers or account_index != 0:
            export_ws["A" + str(header_row)].value = account_index + 1
            export_ws["B" + str(header_row)].value = account.description
            
            accounts_cells = [export_ws[col + str(header_row)] for col in ['A', 'B']]
            for account_cell in accounts_cells:
                FormatCell(account_cell, 1)
            header_row += 1
        
        if full_headers:
            for sub_account in account.sub_accounts:
                export_ws["A" + str(header_row)].value = account.sub_accounts.index(sub_account) + 1
                export_ws["B" + str(header_row)].value = sub_account.description
                
                sub_accounts_cells = [export_ws[col + str(header_row)] for col in ['A', 'B']]
                for sub_account_cell in sub_accounts_cells:
                    FormatCell(sub_account_cell, 2)
                header_row += 1
    
    export_ws['A' + str(header_row)].value = 'Resultado'
    FormatCell(export_ws['A' + str(header_row)], 0)
    export_ws.merge_cells('A' + str(header_row) + ':B' + str(header_row))
    header_row += 1
    
    if full_headers:
        export_ws['A' + str(header_row)].value = 'Estoque - Custo'
        FormatCell(export_ws['A' + str(header_row)], 0)
        export_ws.merge_cells('A' + str(header_row) + ':B' + str(header_row))
    
def Export2XLSX(results:list[MonthResult], export_wb:Workbook) -> None:
    export_ws = export_wb.active
    export_ws.title = export_sheets.geral
    
    GenerateHeaders(export_ws)
    
    result_colum = 3
    for result in results:
        row = 1
        export_ws.cell(row, result_colum).value = '{:02d}-{:04d}'.format(result.month, result.year)
        FormatCell(export_ws.cell(row, result_colum), 0)
        row += 1
        
        for account in result.bills_plan.accounts:
            export_ws.cell(row, result_colum).value = account.value
            FormatCell(export_ws.cell(row, result_colum), 1)
            row += 1
            
            for sub_account in account.sub_accounts:
                export_ws.cell(row, result_colum).value = sub_account.value
                FormatCell(export_ws.cell(row, result_colum), 2)
                row += 1
        
        export_ws.cell(row, result_colum).value = result.result
        export_ws.cell(row+1, result_colum).value = result.stock_cost
        for final in range(row, row+2):
            FormatCell(export_ws.cell(final, result_colum), 0)
        
        result_colum += 1
    
    AutofitWS(export_ws)
        
def ExportParticipation(results:list[MonthResult], export_wb:Workbook) -> None:
    export_wb.create_sheet(export_sheets.participation)
    export_ws = export_wb[export_sheets.participation]
    
    GenerateHeaders(export_ws, False)
    
    result_colum = 3
    for result in results:
        row = 1
        export_ws.cell(row, result_colum).value = '{:02d}-{:04d}'.format(result.month, result.year)
        FormatCell(export_ws.cell(row, result_colum), 0)
        row += 1
        
        for account in result.bills_plan.accounts:
            if result.bills_plan.accounts.index(account) != 0:
                export_ws.cell(row, result_colum).value = account.participation
                FormatCell(export_ws.cell(row, result_colum), 1, True)
                row += 1
        
        export_ws.cell(row, result_colum).value = result.result_percentage
        FormatCell(export_ws.cell(row, result_colum), 0, True)
        
        result_colum += 1
        
    AutofitWS(export_ws)
        
def AutofitWS(ws:Worksheet) -> None:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width